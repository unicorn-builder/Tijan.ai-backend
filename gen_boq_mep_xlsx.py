"""
gen_boq_mep_xlsx.py — BOQ MEP as Excel (.xlsx) — Bilingual FR/EN
Tijan AI — same data as gen_boq_mep_detail.py, Excel output via openpyxl
Columns: N° | Désignation | Unité | Qté | Basic | High-End | Luxury
"""
import io, math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


VERT_FILL = PatternFill(start_color="43A956", end_color="43A956", fill_type="solid")
GRIS_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


# Translation helper
def _t(fr, en, lang='fr'):
    return en if lang == 'en' else fr


def _add_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = VERT_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _add_row(ws, row, values, subtotal=False):
    font = BOLD_FONT if subtotal else NORMAL_FONT
    fill = GRIS_FILL if subtotal else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if isinstance(val, (int, float)) and val != 0:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal='right')


def generer_boq_mep_xlsx(rm, params: dict, lang: str = "fr") -> bytes:
    """Generate BOQ MEP as Excel. Returns xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ MEP"

    d = rm.params
    el = rm.electrique
    pl = rm.plomberie
    cv = rm.cvc
    cf = rm.courants_faibles
    si = rm.securite_incendie
    asc = rm.ascenseurs
    auto = rm.automatisation
    surf = rm.surf_batie_m2

    # Pricing
    try:
        from prix_marche import get_prix_mep
        px = get_prix_mep(d.ville)
    except:
        class _PX:
            transfo_160kva=22_000_000; transfo_250kva=32_000_000; transfo_400kva=48_000_000
            groupe_electrogene_100kva=18_000_000; groupe_electrogene_200kva=32_000_000
            groupe_electrogene_400kva=58_000_000; compteur_monophase=180_000
            compteur_triphase=280_000; canalisation_cuivre_ml=12_000
            luminaire_led_standard=35_000; luminaire_led_premium=85_000
            colonne_montante_ml=22_000; tuyau_pvc_dn100_ml=14_000
            cuve_eau_5000l=850_000; cuve_eau_10000l=1_500_000
            pompe_surpresseur_1kw=450_000; pompe_surpresseur_3kw=850_000
            split_1cv=450_000; split_2cv=750_000; split_cassette_4cv=1_800_000
            vmc_simple_flux=320_000; vmc_double_flux=850_000
            detecteur_fumee=45_000; centrale_incendie_16zones=1_800_000
            cablage_rj45_ml=3_500; prise_rj45=18_000; baie_serveur_12u=850_000
            camera_ip_interieure=180_000; camera_ip_exterieure=280_000
            ascenseur_630kg_r4_r6=28_000_000; ascenseur_630kg_r7_r10=38_000_000
            ascenseur_1000kg_r6_r10=45_000_000
            domotique_logement=850_000; bms_systeme=12_000_000
        px = _PX()

    # Title rows — 7 columns
    ws.merge_cells('A1:G1')
    ws['A1'] = d.nom
    ws['A1'].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells('A2:G2')
    boq_mep_label = _t('BOQ MEP Detaille', 'Detailed MEP BOQ', lang)
    ws['A2'] = f'{boq_mep_label} — {d.ville} — R+{d.nb_niveaux-1} — {d.usage.value.capitalize()}'
    ws['A2'].font = Font(name="Calibri", size=11)
    ws.merge_cells('A3:G3')
    units_label = _t('logements', 'units', lang)
    pers_label = _t('pers.', 'persons', lang)
    ws['A3'] = f'{_t("Surface", "Area", lang)}: {surf:,.0f} m² | {rm.nb_logements} {units_label} | {rm.nb_personnes} {pers_label} | {_t("3 gammes", "3 tiers", lang)}: Basic / High-End / Luxury'
    ws['A3'].font = Font(name="Calibri", size=10, italic=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    dl = _t('FCFA', 'FCFA', lang)
    if lang == 'en':
        headers = ['No.', 'Description', 'Unit', 'Qty',
                   f'Basic ({dl})', f'High-End ({dl})', f'Luxury ({dl})']
    else:
        headers = ['N°', 'Designation', 'Unite', 'Qte',
                   f'Basic ({dl})', f'High-End ({dl})', f'Luxury ({dl})']
    row = 5
    _add_header_row(ws, row, headers)
    row += 1

    totals_b, totals_h, totals_l = 0, 0, 0

    # ── LOT E — Electrical ──────────────────────────────────────────
    transfo_pu = {100: px.transfo_160kva, 160: px.transfo_160kva,
                  250: px.transfo_250kva, 400: px.transfo_400kva}.get(el.transfo_kva, px.transfo_160kva)
    ge_pu = {100: px.groupe_electrogene_100kva, 200: px.groupe_electrogene_200kva,
             400: px.groupe_electrogene_400kva}.get(el.groupe_electrogene_kva, px.groupe_electrogene_200kva)
    L_cable = int(surf * 2.5)
    nb_lum = int(surf / 12)

    transfo_label = _t('Transformateur', 'Transformer', lang)
    tgbt_label = _t('Tableau general basse tension', 'LVMD', lang)
    genset_label = _t('Groupe electrogene', 'Diesel generator', lang)
    meters_label = _t('Compteurs', 'Meters', lang)
    cable_label = _t('Cablage cuivre', 'Copper cabling', lang)
    lights_label = _t('Luminaires LED', 'LED luminaires', lang)

    _add_row(ws, row, ('E.1', f'{transfo_label} {el.transfo_kva} kVA', 'U', 1,
                       transfo_pu, int(transfo_pu*1.15), int(transfo_pu*1.40)))
    row += 1
    _add_row(ws, row, ('E.2', tgbt_label, 'U', 1,
                       3_500_000, 5_500_000, 8_000_000))
    row += 1
    _add_row(ws, row, ('E.3', f'{genset_label} {el.groupe_electrogene_kva} kVA', 'U', 1,
                       ge_pu, int(ge_pu*1.20), int(ge_pu*1.45)))
    row += 1
    _add_row(ws, row, ('E.4', f'{meters_label} ({el.nb_compteurs})', 'U', el.nb_compteurs,
                       px.compteur_monophase, px.compteur_triphase, int(px.compteur_triphase*1.3)))
    row += 1
    _add_row(ws, row, ('E.5', f'{cable_label} {L_cable} ml', 'ml', L_cable,
                       px.canalisation_cuivre_ml, int(px.canalisation_cuivre_ml*1.3), int(px.canalisation_cuivre_ml*1.6)))
    row += 1
    _add_row(ws, row, ('E.6', f'{lights_label} ({nb_lum})', 'U', nb_lum,
                       px.luminaire_led_standard, px.luminaire_led_premium, int(px.luminaire_led_premium*1.5)))
    row += 1

    c_b = transfo_pu + 3_500_000 + ge_pu + el.nb_compteurs*px.compteur_monophase + L_cable*px.canalisation_cuivre_ml + nb_lum*px.luminaire_led_standard
    c_h, c_l = int(c_b*1.35), int(c_b*1.75)
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT E', 'SUBTOTAL LOT E', lang), '', '', c_b, c_h, c_l), subtotal=True)
    totals_b += c_b; totals_h += c_h; totals_l += c_l
    row += 1

    # ── LOT P — Plumbing ────────────────────────────────────────────
    cuve_pu = px.cuve_eau_5000l if pl.volume_citerne_m3 <= 5 else px.cuve_eau_10000l
    surp_pu = px.pompe_surpresseur_1kw if pl.debit_surpresseur_m3h <= 4 else px.pompe_surpresseur_3kw
    cistern_label = _t('Citerne', 'Water cistern', lang)
    pump_label = _t('Surpresseur', 'Booster pump', lang)
    risers_label = _t('Colonnes montantes', 'Vertical stacks', lang)
    network_label = _t('Reseau EU/EV PVC', 'Fresh water/waste water PVC', lang)

    nb_cuves = max(1, math.ceil(pl.volume_citerne_m3/10))
    ml_risers = int(d.nb_niveaux*d.hauteur_etage_m*1.2)
    ml_pvc = int(surf*0.15)

    _add_row(ws, row, ('P.1', f'{cistern_label} {int(pl.volume_citerne_m3)} m³', 'U', nb_cuves,
                       cuve_pu, int(cuve_pu*1.2), int(cuve_pu*1.5)))
    row += 1
    _add_row(ws, row, ('P.2', f'{pump_label} {pl.debit_surpresseur_m3h} m³/h', 'U', 1,
                       surp_pu, int(surp_pu*1.3), int(surp_pu*1.7)))
    row += 1
    _add_row(ws, row, ('P.3', f'{risers_label} DN{pl.diam_colonne_montante_mm}', 'ml', ml_risers,
                       px.colonne_montante_ml, int(px.colonne_montante_ml*1.2), int(px.colonne_montante_ml*1.5)))
    row += 1
    _add_row(ws, row, ('P.4', f'{network_label} DN100', 'ml', ml_pvc,
                       px.tuyau_pvc_dn100_ml, int(px.tuyau_pvc_dn100_ml*1.2), int(px.tuyau_pvc_dn100_ml*1.5)))
    row += 1

    c_pb = cuve_pu + surp_pu + ml_risers*px.colonne_montante_ml + ml_pvc*px.tuyau_pvc_dn100_ml
    c_ph, c_pl = int(c_pb*1.25), int(c_pb*1.60)
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT P', 'SUBTOTAL LOT P', lang), '', '', c_pb, c_ph, c_pl), subtotal=True)
    totals_b += c_pb; totals_h += c_ph; totals_l += c_pl
    row += 1

    # ── LOT C — HVAC ────────────────────────────────────────────────
    nb_splits = int(surf / 25)
    split_pu = px.split_1cv if surf < 1000 else px.split_2cv
    nb_vmc = int(surf / 50)
    ac_label = _t('Climatiseurs split', 'Split AC units', lang)
    vmc_label = _t('VMC / Ventilation', 'HVAC / Ventilation', lang)

    _add_row(ws, row, ('C.1', f'{ac_label} ({nb_splits})', 'U', nb_splits,
                       split_pu, int(split_pu*1.4), int(split_pu*2.0)))
    row += 1
    _add_row(ws, row, ('C.2', vmc_label, 'U', nb_vmc,
                       px.vmc_simple_flux, px.vmc_double_flux, int(px.vmc_double_flux*1.5)))
    row += 1

    c_cvc_b = nb_splits * split_pu + nb_vmc * px.vmc_simple_flux
    c_cvc_h, c_cvc_l = int(c_cvc_b*1.40), int(c_cvc_b*2.0)
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT C', 'SUBTOTAL LOT C', lang), '', '', c_cvc_b, c_cvc_h, c_cvc_l), subtotal=True)
    totals_b += c_cvc_b; totals_h += c_cvc_h; totals_l += c_cvc_l
    row += 1

    # ── LOT SSI — Fire Safety ───────────────────────────────────────
    nb_det = int(surf / 30)
    smoke_label = _t('Detecteurs fumee', 'Smoke detectors', lang)
    fire_panel_label = _t('Centrale incendie', 'Fire control panel', lang)

    _add_row(ws, row, ('SSI.1', f'{smoke_label} ({nb_det})', 'U', nb_det,
                       px.detecteur_fumee, int(px.detecteur_fumee*1.5), int(px.detecteur_fumee*2.2)))
    row += 1
    _add_row(ws, row, ('SSI.2', fire_panel_label, 'U', 1,
                       px.centrale_incendie_16zones, int(px.centrale_incendie_16zones*1.5), int(px.centrale_incendie_16zones*2.0)))
    row += 1

    c_ssi_b = nb_det * px.detecteur_fumee + px.centrale_incendie_16zones
    c_ssi_h, c_ssi_l = int(c_ssi_b*1.5), int(c_ssi_b*2.2)
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT SSI', 'SUBTOTAL LOT SSI', lang), '', '', c_ssi_b, c_ssi_h, c_ssi_l), subtotal=True)
    totals_b += c_ssi_b; totals_h += c_ssi_h; totals_l += c_ssi_l
    row += 1

    # ── LOT CF — Low Voltage ────────────────────────────────────────
    nb_rj = int(surf / 15)
    rj45_label = _t('Prises RJ45', 'RJ45 outlets', lang)
    cabling_label = _t('Cablage RJ45', 'RJ45 cabling', lang)

    _add_row(ws, row, ('CF.1', f'{rj45_label} ({nb_rj})', 'U', nb_rj,
                       px.prise_rj45, int(px.prise_rj45*1.5), int(px.prise_rj45*2.0)))
    row += 1
    _add_row(ws, row, ('CF.2', cabling_label, 'ml', nb_rj*15,
                       px.cablage_rj45_ml, int(px.cablage_rj45_ml*1.3), int(px.cablage_rj45_ml*1.6)))
    row += 1

    c_cf_b = nb_rj * (px.cablage_rj45_ml * 15 + px.prise_rj45) + px.baie_serveur_12u
    c_cf_h, c_cf_l = int(c_cf_b*1.4), int(c_cf_b*2.0)
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT CF', 'SUBTOTAL LOT CF', lang), '', '', c_cf_b, c_cf_h, c_cf_l), subtotal=True)
    totals_b += c_cf_b; totals_h += c_cf_h; totals_l += c_cf_l
    row += 2

    # ── GRAND TOTAL ─────────────────────────────────────────────────
    total_label = _t('TOTAL GENERAL MEP', 'TOTAL MEP COST', lang)
    _add_row(ws, row, ('', total_label, '', '', totals_b, totals_h, totals_l), subtotal=True)
    for col in range(5, 8):
        ws.cell(row=row, column=col).font = Font(name="Calibri", size=12, bold=True, color="43A956")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
