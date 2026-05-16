"""
gen_boq_xlsx.py — BOQ Structure as Excel (.xlsx) — Bilingual FR/EN
Tijan AI — same data as gen_boq_structure.py, Excel output via openpyxl
Format BPU : N° | Désignation | Unité | Qté | PU | Montant | Observations
"""
import io, math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers


VERT_FILL = PatternFill(start_color="43A956", end_color="43A956", fill_type="solid")
GRIS_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


# Translation helpers
def _t(fr, en, lang='fr'):
    return en if lang == 'en' else fr


def _pu(montant, qte):
    return int(montant / qte) if qte and qte > 0 else 0


def _add_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = VERT_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _add_row(ws, row, values, bold=False, subtotal=False):
    font = BOLD_FONT if bold or subtotal else NORMAL_FONT
    fill = GRIS_FILL if subtotal else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if isinstance(val, (int, float)) and val != 0:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal='right')


def generer_boq_structure_xlsx(rs, params: dict, lang: str = "fr") -> bytes:
    """Generate BOQ Structure as Excel. Returns xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Structure"

    d = rs.params
    boq = rs.boq
    poteaux = rs.poteaux
    poutre = rs.poutre_principale
    dalle = rs.dalle
    fond = rs.fondation

    # Pricing
    try:
        from prix_marche import get_prix_structure
        px = get_prix_structure(d.ville)
    except:
        class _PX:
            beton_c2530_m3=170000; beton_c3037_m3=185000; beton_c3545_m3=210000
            acier_ha400_kg=750; acier_ha500_kg=810; coffrage_bois_m2=18000
            coffrage_metal_m2=25000; terr_mecanique_m3=8500; terr_manuel_m3=5000
            remblai_m3=6500; pieu_fore_d600_ml=220000; pieu_fore_d800_ml=285000
            pieu_fore_d1000_ml=360000; semelle_filante_ml=85000; radier_m2=95000
            agglo_creux_10_m2=18000; agglo_creux_15_m2=24000; agglo_creux_20_m2=30000
            mo_chef_chantier_j=35000; mo_macon_j=18000; mo_ferrailleur_j=20000
            mo_manœuvre_j=8000
        px = _PX()

    prix_beton = {
        'C25/30': px.beton_c2530_m3, 'C30/37': px.beton_c3037_m3,
        'C35/45': px.beton_c3545_m3, 'C40/50': getattr(px, 'beton_c4050_m3', px.beton_c3545_m3),
    }.get(rs.classe_beton, px.beton_c3037_m3)
    prix_acier = px.acier_ha400_kg if rs.classe_acier == 'HA400' else px.acier_ha500_kg

    nb_pot = (d.nb_travees_x + 1) * (d.nb_travees_y + 1)
    surf_batie = boq.surface_batie_m2

    # Title rows
    ws.merge_cells('A1:G1')
    ws['A1'] = d.nom
    ws['A1'].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells('A2:G2')
    ws['A2'] = f'BOQ Structure — {d.ville} — R+{d.nb_niveaux-1} — {d.usage.value.capitalize()}'
    ws['A2'].font = Font(name="Calibri", size=11)
    ws.merge_cells('A3:G3')
    built_area_label = _t('Surface bâtie', 'Built area', lang)
    ws['A3'] = f'{built_area_label}: {surf_batie:,.0f} m² — {_t("Béton", "Concrete", lang)} {rs.classe_beton} — {_t("Acier", "Steel", lang)} {rs.classe_acier} — {_t("Prix 2026", "Pricing 2026", lang)}'
    ws['A3'].font = Font(name="Calibri", size=10, italic=True)

    # Column widths — 7 columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30

    if lang == 'en':
        headers = ['No.', 'Description', 'Unit', 'Qty', 'Unit Price (FCFA)', 'Amount (FCFA)', 'Notes']
    else:
        headers = ['N°', 'Désignation', 'Unité', 'Qté', 'PU (FCFA)', 'Montant (FCFA)', 'Observations']
    row = 5
    _add_header_row(ws, row, headers)
    row += 1

    grand_total = 0

    # LOT 1 — Installation
    q_cloture = int(4*math.sqrt(d.surface_emprise_m2))
    m_cloture = int(q_cloture*15000)
    m_base = int(surf_batie*800)
    m_branch = int(surf_batie*500)
    m_secu = int(surf_batie*300)
    m_repli = int(surf_batie*600)
    inst_forfait = m_cloture + m_base + m_branch + m_secu + m_repli

    lot1_rows = [
        ('1.1', _t('Clôture de chantier', 'Site fence', lang), 'ml', q_cloture, 15000, m_cloture, ''),
        ('1.2', _t('Base vie chantier', 'Temporary site facilities', lang), _t('forfait', 'lump sum', lang), 1, m_base, m_base, _t('Modulaires', 'Modular units', lang)),
        ('1.3', _t('Branchements provisoires eau + élec', 'Temporary water + power connections', lang), _t('forfait', 'lump sum', lang), 1, m_branch, m_branch, ''),
        ('1.4', _t('Signalétique sécurité + EPI', 'Safety signage + PPE', lang), _t('forfait', 'lump sum', lang), 1, m_secu, m_secu, ''),
        ('1.5', _t('Repli et nettoyage', 'Demobilization & cleanup', lang), _t('forfait', 'lump sum', lang), 1, m_repli, m_repli, ''),
    ]
    for vals in lot1_rows:
        _add_row(ws, row, vals)
        row += 1
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 1', 'SUBTOTAL LOT 1', lang), '', '', '', inst_forfait, ''), subtotal=True)
    grand_total += inst_forfait
    row += 1

    # LOT 2 — Terrassement / Earthwork
    V_decap = d.surface_emprise_m2 * 0.30
    V_fouilles = d.surface_emprise_m2 * (fond.profondeur_m + 0.50)
    V_remblai = V_fouilles * 0.30
    V_evacu = V_fouilles * 0.70

    m_decap = int(V_decap*px.terr_mecanique_m3)
    m_fouilles = int(V_fouilles*px.terr_mecanique_m3)
    m_remblai = int(V_remblai*px.remblai_m3)
    m_evacu = int(V_evacu*5000)
    c_terr = m_decap + m_fouilles + m_remblai + m_evacu

    lot2_rows = [
        ('2.1', _t('Décapage terre végétale e=30cm', 'Topsoil removal d=30cm', lang), 'm³', int(V_decap), px.terr_mecanique_m3, m_decap, ''),
        ('2.2', _t('Fouilles générales mécaniques', 'Mechanical excavation', lang), 'm³', int(V_fouilles), px.terr_mecanique_m3, m_fouilles, ''),
        ('2.3', _t('Remblai compacté', 'Compacted backfill', lang), 'm³', int(V_remblai), px.remblai_m3, m_remblai, ''),
        ('2.4', _t('Évacuation terres', 'Soil evacuation', lang), 'm³', int(V_evacu), 5000, m_evacu, ''),
    ]
    for vals in lot2_rows:
        _add_row(ws, row, vals)
        row += 1
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 2', 'SUBTOTAL LOT 2', lang), '', '', '', c_terr, ''), subtotal=True)
    grand_total += c_terr
    row += 1

    # LOT 3 — Fondations / Foundations
    if fond.nb_pieux > 0:
        prix_pieu = {600: px.pieu_fore_d600_ml, 800: px.pieu_fore_d800_ml,
                     1000: px.pieu_fore_d1000_ml}.get(fond.diam_pieu_mm, px.pieu_fore_d800_ml)
        q_pieux_ml = int(fond.nb_pieux * fond.longueur_pieu_m)
        c_pieux = int(q_pieux_ml * prix_pieu)
        q_longr = int(nb_pot * 6)
        c_longr = int(q_longr * 85000)
        c_fond = c_pieux + c_longr
        piles_label = _t('pieux', 'piles', lang)
        lot3_rows = [
            ('3.1', _t(f'Pieux forés Ø{fond.diam_pieu_mm}mm L={fond.longueur_pieu_m}m', f'Bored piles Ø{fond.diam_pieu_mm}mm L={fond.longueur_pieu_m}m', lang), 'ml', q_pieux_ml, prix_pieu, c_pieux, f'{fond.nb_pieux} {piles_label}'),
            ('3.2', _t('Longrines BA 30×50cm', 'Concrete pile caps 30×50cm', lang), 'ml', q_longr, 85000, c_longr, ''),
        ]
    else:
        q_proprete = int(d.surface_emprise_m2*0.10)
        m_proprete = int(q_proprete*120000) if q_proprete > 0 else int(d.surface_emprise_m2*0.10*120000)
        q_sem = int(fond.beton_semelle_m3*0.6)
        m_semelles = int(q_sem*prix_beton) if q_sem > 0 else int(fond.beton_semelle_m3*0.6*prix_beton)
        q_arm = int(fond.beton_semelle_m3*100)
        m_arm_sem = int(q_arm*prix_acier)
        c_fond = m_proprete + m_semelles + m_arm_sem
        lot3_rows = [
            ('3.1', _t('Béton de propreté e=10cm', 'Blinding concrete d=10cm', lang), 'm³', max(1, q_proprete), _pu(m_proprete, max(1, q_proprete)), m_proprete, ''),
            ('3.2', _t(f'Semelles BA {rs.classe_beton}', f'Concrete footings {rs.classe_beton}', lang), 'm³', max(1, q_sem), prix_beton, m_semelles, ''),
            ('3.3', _t(f'Armatures semelles {rs.classe_acier}', f'Footing reinforcement {rs.classe_acier}', lang), 'kg', q_arm, prix_acier, m_arm_sem, ''),
        ]
    for vals in lot3_rows:
        _add_row(ws, row, vals)
        row += 1
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 3', 'SUBTOTAL LOT 3', lang), '', '', '', c_fond, ''), subtotal=True)
    grand_total += c_fond
    row += 1

    # LOT 4 — Structure BA / Reinforced Concrete Structure
    ep_dalle_m = dalle.epaisseur_mm / 1000
    c_struct = 0

    # 4.1 — Poteaux / Columns (per level)
    for i, pt in enumerate(poteaux):
        b = pt.section_mm / 1000
        V_niv = b**2 * d.hauteur_etage_m * nb_pot
        As_niv = pt.nb_barres * math.pi * pt.diametre_mm**2 / 400 * nb_pot * d.hauteur_etage_m * 7850 / 10000
        c_b = int(V_niv * prix_beton)
        c_a = int(As_niv * prix_acier)
        c_niv = c_b + c_a
        cols_desc = _t(f'Poteaux {pt.niveau} — {pt.section_mm}×{pt.section_mm}mm — {pt.nb_barres}HA{pt.diametre_mm}',
                       f'Columns {pt.niveau} — {pt.section_mm}×{pt.section_mm}mm — {pt.nb_barres}HA{pt.diametre_mm}', lang)
        _add_row(ws, row, (f'4.1.{i+1}', cols_desc, 'U', nb_pot, _pu(c_niv, nb_pot), c_niv, f'NEd={pt.NEd_kN:.0f}kN'))
        row += 1
        c_struct += c_niv

    # 4.2 — Poutres / Beams
    pp_b = poutre.b_mm / 1000
    pp_h = poutre.h_mm / 1000
    L_px = d.nb_travees_x * d.portee_max_m * (d.nb_travees_y+1) * d.nb_niveaux
    L_py = d.nb_travees_y * d.portee_min_m * (d.nb_travees_x+1) * d.nb_niveaux
    V_pout = (L_px + L_py) * pp_b * (pp_h - ep_dalle_m)
    kg_pout = V_pout * 100
    c_pout_beton = int(V_pout * prix_beton)
    c_pout_acier = int(kg_pout * prix_acier)
    beams_desc = _t(f'Béton poutres {poutre.b_mm}×{poutre.h_mm}mm',
                    f'Beam concrete {poutre.b_mm}×{poutre.h_mm}mm', lang)
    _add_row(ws, row, ('4.2.1', beams_desc, 'm³', f'{V_pout:.1f}', prix_beton, c_pout_beton, ''))
    row += 1
    beams_steel = _t('Armatures poutres', 'Beam reinforcement', lang)
    _add_row(ws, row, ('4.2.2', beams_steel, 'kg', f'{kg_pout:.0f}', prix_acier, c_pout_acier, f'As={poutre.As_inf_cm2}cm²'))
    row += 1
    c_struct += c_pout_beton + c_pout_acier

    # 4.3 — Dalle / Slab
    V_dalle = ep_dalle_m * surf_batie * 0.85
    c_dalle = int(V_dalle * prix_beton)
    kg_dalle = V_dalle * 80
    c_dalle_a = int(kg_dalle * prix_acier)
    slab_desc = _t(f'Béton dalles ep.{dalle.epaisseur_mm}mm', f'Slab concrete t={dalle.epaisseur_mm}mm', lang)
    _add_row(ws, row, ('4.3.1', slab_desc, 'm³', f'{V_dalle:.1f}', prix_beton, c_dalle, ''))
    row += 1
    slab_steel = _t('Armatures dalle HA10 deux sens', 'Slab reinforcement HA10 both ways', lang)
    _add_row(ws, row, ('4.3.2', slab_steel, 'kg', f'{kg_dalle:.0f}', prix_acier, c_dalle_a, ''))
    row += 1
    c_struct += c_dalle + c_dalle_a

    # 4.4 — Coffrage / Formwork
    coff_total = int(surf_batie * 0.85)
    c_coff = int(coff_total * px.coffrage_bois_m2)
    coff_desc = _t('Coffrage (poteaux + poutres + dalle)', 'Formwork (columns + beams + slab)', lang)
    _add_row(ws, row, ('4.4', coff_desc, 'm²', coff_total, px.coffrage_bois_m2, c_coff, ''))
    row += 1
    c_struct += c_coff

    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 4', 'SUBTOTAL LOT 4', lang), '', '', '', c_struct, ''), subtotal=True)
    grand_total += c_struct
    row += 2

    # GRAND TOTAL
    total_label = _t('TOTAL GÉNÉRAL STRUCTURE', 'TOTAL STRUCTURAL COST', lang)
    unit_label = 'FCFA/m²'
    _add_row(ws, row, ('', total_label, '', '', '', grand_total, f'{grand_total/surf_batie:,.0f} {unit_label}'), bold=True)
    ws.cell(row=row, column=6).font = Font(name="Calibri", size=12, bold=True, color="43A956")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
