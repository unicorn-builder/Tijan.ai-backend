"""
gen_planning_xlsx.py — Excel planning + cash flow generator
Tijan AI — Gantt chart, monthly cash flow, lot×phase cross-tabulation
Uses openpyxl with charts (BarChart, LineChart) for visualizations.
"""
import io
from dataclasses import dataclass, field
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styles — matching gen_boq_xlsx.py
# ---------------------------------------------------------------------------
VERT_FILL = PatternFill(start_color="43A956", end_color="43A956", fill_type="solid")
GRIS_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
BLEU_FILL = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11)
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Gantt bar colours per lot category
_LOT_CATEGORY = {
    "Structure": VERT_FILL,
    "MEP": BLEU_FILL,
    "Finitions": ORANGE_FILL,
}

# Light versions for Gantt cells
VERT_LIGHT = PatternFill(start_color="A8E6B4", end_color="A8E6B4", fill_type="solid")
BLEU_LIGHT = PatternFill(start_color="A8CBE6", end_color="A8CBE6", fill_type="solid")
ORANGE_LIGHT = PatternFill(start_color="F5CBA7", end_color="F5CBA7", fill_type="solid")

_LOT_CATEGORY_LIGHT = {
    "Structure": VERT_LIGHT,
    "MEP": BLEU_LIGHT,
    "Finitions": ORANGE_LIGHT,
}

# ---------------------------------------------------------------------------
# Translation helper
# ---------------------------------------------------------------------------
def _t(fr: str, en: str, lang: str = "fr") -> str:
    return en if lang == "en" else fr


# ---------------------------------------------------------------------------
# Helper: write a header row
# ---------------------------------------------------------------------------
def _add_header_row(ws, row: int, headers: list, fill=None):
    """Write a row of header cells with HEADER_FONT styling."""
    fill = fill or VERT_FILL
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _add_row(ws, row: int, values: list, bold: bool = False, subtotal: bool = False):
    """Write a data row with optional bold/subtotal styling."""
    font = BOLD_FONT if bold or subtotal else NORMAL_FONT
    fill = GRIS_FILL if subtotal else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if isinstance(val, (int, float)) and val != 0:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")


def _set_cell(ws, row: int, col: int, value, font=None, fill=None,
              fmt=None, alignment=None, border=None):
    """Set a single cell with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


# ---------------------------------------------------------------------------
# Determine lot category for Gantt colour
# ---------------------------------------------------------------------------
_STRUCTURE_LOTS = {
    "Installation chantier", "Terrassement", "Fondations",
    "Structure BA", "Maçonnerie", "Étanchéité",
    "Site installation", "Earthwork", "Foundations",
    "RC Structure", "Masonry", "Waterproofing",
}
_MEP_LOTS = {
    "Plomberie", "Électricité", "CVC", "HVAC",
    "Sécurité incendie", "Courants faibles",
    "Plumbing", "Electrical", "Fire safety", "Low current",
}


def _lot_category(lot_name: str) -> str:
    """Return 'Structure', 'MEP', or 'Finitions' based on lot name."""
    if lot_name in _STRUCTURE_LOTS:
        return "Structure"
    if lot_name in _MEP_LOTS:
        return "MEP"
    return "Finitions"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def generer_planning_xlsx(planning, lang: str = "fr") -> bytes:
    """Generate Excel planning + cash flow workbook.

    Args:
        planning: A Planning object with attributes:
            - taches: list of task objects (id, lot, designation, duree_jours,
              debut_jour, fin_jour, mois_debut, mois_fin,
              cout_materiaux, cout_pose, cout_total, phase, categorie)
            - phases: list of phase names
            - duree_totale_mois: total project duration in months
            - tresorerie_mensuelle: list of dicts with keys
              mois, materiaux, pose, total, cumul
            - tresorerie_lot_phase: list of dicts with keys
              lot, phase, montant
            - nom_projet: project name (optional)
            - cout_total: total project cost
        lang: "fr" or "en"

    Returns:
        bytes of the .xlsx file
    """
    wb = Workbook()

    # Gather data
    taches = planning.taches
    phases = getattr(planning, "phases", [])
    duree_mois = getattr(planning, "duree_totale_mois", 0)
    treso_mens = getattr(planning, "tresorerie_mensuelle", [])
    treso_lot = getattr(planning, "tresorerie_lot_phase", [])
    nom_projet = getattr(planning, "nom_projet", "Projet")
    cout_total_projet = getattr(planning, "cout_total", 0)

    # Compute total months from tasks if not provided
    if duree_mois == 0 and taches:
        duree_mois = max(t.mois_fin for t in taches)

    # Month labels
    mois_labels = [
        _t(f"Mois {m}", f"Month {m}", lang) for m in range(1, duree_mois + 1)
    ]

    # -----------------------------------------------------------------------
    # Sheet 1: Planning (Gantt)
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Planning"

    # Title
    ws1.merge_cells("A1:K1")
    _set_cell(ws1, 1, 1, _t(
        f"Planning d'exécution — {nom_projet}",
        f"Execution schedule — {nom_projet}", lang
    ), font=TITLE_FONT)

    # Headers
    base_headers = [
        _t("N°", "ID", lang),
        _t("Lot", "Lot", lang),
        _t("Désignation", "Description", lang),
        _t("Durée (j)", "Duration (d)", lang),
        _t("Début (j)", "Start (d)", lang),
        _t("Fin (j)", "End (d)", lang),
        _t("Mois début", "Start month", lang),
        _t("Mois fin", "End month", lang),
        _t("Coût matériaux", "Material cost", lang),
        _t("Coût pose", "Labour cost", lang),
        _t("Coût total", "Total cost", lang),
    ]
    all_headers = base_headers + mois_labels

    header_row = 3
    _add_header_row(ws1, header_row, all_headers)

    # Column widths
    col_widths = [6, 20, 35, 10, 10, 10, 10, 10, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    # Month columns narrow
    for i in range(12, 12 + duree_mois):
        ws1.column_dimensions[get_column_letter(i)].width = 5

    # Freeze panes: row 4 (below header), column L (12)
    ws1.freeze_panes = "L4"

    # Group tasks by lot
    lots_ordered = []
    lots_tasks = {}
    for t in taches:
        lot = t.lot
        if lot not in lots_tasks:
            lots_ordered.append(lot)
            lots_tasks[lot] = []
        lots_tasks[lot].append(t)

    data_row = header_row + 1
    total_materiaux = 0
    total_pose = 0
    total_cout = 0

    for lot in lots_ordered:
        tasks_in_lot = lots_tasks[lot]
        cat = _lot_category(lot)
        cat_fill = _LOT_CATEGORY.get(cat, VERT_FILL)
        cat_fill_light = _LOT_CATEGORY_LIGHT.get(cat, VERT_LIGHT)

        # Lot header row
        lot_mat = sum(getattr(t, "cout_materiaux", 0) for t in tasks_in_lot)
        lot_pose = sum(getattr(t, "cout_pose", 0) for t in tasks_in_lot)
        lot_total = sum(getattr(t, "cout_total", 0) for t in tasks_in_lot)

        _add_row(ws1, data_row, [
            "", lot, "", "", "", "", "", "",
            lot_mat, lot_pose, lot_total,
        ], bold=True, subtotal=True)
        # Colour the lot header row Gantt area
        lot_mois_min = min(t.mois_debut for t in tasks_in_lot)
        lot_mois_max = max(t.mois_fin for t in tasks_in_lot)
        for m in range(lot_mois_min, lot_mois_max + 1):
            col_idx = 12 + m - 1  # column 12 = month 1
            if 1 <= col_idx <= 12 + duree_mois:
                _set_cell(ws1, data_row, col_idx, "", fill=cat_fill,
                          border=THIN_BORDER)
        data_row += 1

        # Task rows
        for t in tasks_in_lot:
            values = [
                getattr(t, "id", ""),
                lot,
                getattr(t, "designation", ""),
                getattr(t, "duree_jours", 0),
                getattr(t, "debut_jour", 0),
                getattr(t, "fin_jour", 0),
                getattr(t, "mois_debut", 0),
                getattr(t, "mois_fin", 0),
                getattr(t, "cout_materiaux", 0),
                getattr(t, "cout_pose", 0),
                getattr(t, "cout_total", 0),
            ]
            _add_row(ws1, data_row, values)

            # Gantt bars
            m_start = getattr(t, "mois_debut", 0)
            m_end = getattr(t, "mois_fin", 0)
            for m in range(m_start, m_end + 1):
                col_idx = 12 + m - 1
                if 1 <= col_idx <= 12 + duree_mois:
                    _set_cell(ws1, data_row, col_idx, "", fill=cat_fill_light,
                              border=THIN_BORDER)

            total_materiaux += getattr(t, "cout_materiaux", 0)
            total_pose += getattr(t, "cout_pose", 0)
            total_cout += getattr(t, "cout_total", 0)
            data_row += 1

    # Totals row
    data_row += 1
    total_values = [
        "", _t("TOTAL", "TOTAL", lang), "", "", "", "", "", "",
        total_materiaux, total_pose, total_cout,
    ]
    _add_row(ws1, data_row, total_values, bold=True, subtotal=True)
    ws1.cell(row=data_row, column=11).font = Font(
        name="Calibri", size=12, bold=True, color="43A956"
    )

    # -----------------------------------------------------------------------
    # Sheet 2: Trésorerie mensuelle
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet(_t("Trésorerie mensuelle", "Monthly cash flow", lang))

    # Title
    ws2.merge_cells("A1:F1")
    _set_cell(ws2, 1, 1, _t(
        f"Trésorerie mensuelle — {nom_projet}",
        f"Monthly cash flow — {nom_projet}", lang
    ), font=TITLE_FONT)

    treso_headers = [
        _t("Mois", "Month", lang),
        _t("Matériaux / Équipements", "Materials / Equipment", lang),
        _t("Main d'œuvre / Pose", "Labour / Installation", lang),
        _t("Total mensuel", "Monthly total", lang),
        _t("Cumul", "Cumulative", lang),
        _t("% du total", "% of total", lang),
    ]
    _add_header_row(ws2, 3, treso_headers)

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 12

    ws2.freeze_panes = "A4"

    # Compute grand total for percentages
    grand_total = cout_total_projet
    if grand_total == 0 and treso_mens:
        grand_total = sum(m.get("total", 0) if isinstance(m, dict)
                         else getattr(m, "total", 0) for m in treso_mens)
    if grand_total == 0:
        grand_total = 1  # avoid division by zero

    row = 4
    cumul = 0
    sum_mat = 0
    sum_pose = 0
    sum_total = 0

    for entry in treso_mens:
        if isinstance(entry, dict):
            mois = entry.get("mois", row - 3)
            mat = entry.get("materiaux", 0)
            pose = entry.get("pose", 0)
            total = entry.get("total", 0)
            cum = entry.get("cumul", 0)
        else:
            mois = getattr(entry, "mois", row - 3)
            mat = getattr(entry, "materiaux", 0)
            pose = getattr(entry, "pose", 0)
            total = getattr(entry, "total", 0)
            cum = getattr(entry, "cumul", 0)

        if cum == 0:
            cumul += total
            cum = cumul

        pct = cum / grand_total if grand_total else 0

        mois_label = _t(f"Mois {mois}", f"Month {mois}", lang)

        for col, val in enumerate([mois_label, mat, pose, total, cum, pct], 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col == 6:
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col >= 2:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")

        sum_mat += mat
        sum_pose += pose
        sum_total += total
        row += 1

    # Totals row
    total_row = row
    _add_row(ws2, total_row, [
        _t("TOTAL", "TOTAL", lang), sum_mat, sum_pose, sum_total,
        sum_total, 1.0,
    ], bold=True, subtotal=True)
    ws2.cell(row=total_row, column=6).number_format = PCT_FMT

    # --- Bar chart: monthly costs ---
    nb_months = len(treso_mens)
    if nb_months > 0:
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = _t("Coûts mensuels (FCFA)", "Monthly costs (FCFA)", lang)
        chart1.y_axis.title = "FCFA"
        chart1.x_axis.title = _t("Mois", "Month", lang)
        chart1.style = 10
        chart1.width = 28
        chart1.height = 14

        # Data: columns B (materiaux), C (pose), D (total)
        data_ref = Reference(ws2, min_col=2, max_col=4,
                             min_row=3, max_row=3 + nb_months)
        cats_ref = Reference(ws2, min_col=1,
                             min_row=4, max_row=3 + nb_months)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.shape = 4

        # --- Line chart overlay: cumulative S-curve ---
        chart2 = LineChart()
        chart2.title = _t("Cumul (courbe en S)", "Cumulative (S-curve)", lang)
        chart2.y_axis.axId = 200
        chart2.y_axis.title = _t("Cumul FCFA", "Cumulative FCFA", lang)
        chart2.y_axis.crosses = "max"
        chart2.style = 10

        cumul_ref = Reference(ws2, min_col=5,
                              min_row=3, max_row=3 + nb_months)
        chart2.add_data(cumul_ref, titles_from_data=True)
        chart2.set_categories(cats_ref)

        # Style line
        s = chart2.series[0]
        s.graphicalProperties.line.width = 25000  # ~2pt in EMU

        # Combine
        chart1 += chart2

        ws2.add_chart(chart1, f"A{total_row + 3}")

    # -----------------------------------------------------------------------
    # Sheet 3: Trésorerie par lot
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet(_t("Trésorerie par lot", "Cash flow by lot", lang))

    ws3.merge_cells("A1:F1")
    _set_cell(ws3, 1, 1, _t(
        f"Trésorerie par lot et par phase — {nom_projet}",
        f"Cash flow by lot and phase — {nom_projet}", lang
    ), font=TITLE_FONT)

    # Build cross-tabulation data
    # Gather unique lots and phases from data
    lot_order = []
    lot_set = set()
    phase_order = phases if phases else []
    phase_set = set(phases) if phases else set()

    # Index: (lot, phase) -> montant
    cross = {}
    for entry in treso_lot:
        if isinstance(entry, dict):
            lot = entry.get("lot", "")
            phase = entry.get("phase", "")
            montant = entry.get("montant", 0)
        else:
            lot = getattr(entry, "lot", "")
            phase = getattr(entry, "phase", "")
            montant = getattr(entry, "montant", 0)

        if lot not in lot_set:
            lot_set.add(lot)
            lot_order.append(lot)
        if phase not in phase_set:
            phase_set.add(phase)
            phase_order.append(phase)

        cross[(lot, phase)] = cross.get((lot, phase), 0) + montant

    nb_phases = len(phase_order)

    # Headers: Lot | Phase1 | Phase2 | ... | Total | %
    lot_headers = [_t("Lot", "Lot", lang)] + phase_order + [
        _t("Total", "Total", lang),
        _t("% du total", "% of total", lang),
    ]
    _add_header_row(ws3, 3, lot_headers)

    ws3.column_dimensions["A"].width = 22
    for i in range(2, 2 + nb_phases):
        ws3.column_dimensions[get_column_letter(i)].width = 18
    ws3.column_dimensions[get_column_letter(2 + nb_phases)].width = 18
    ws3.column_dimensions[get_column_letter(3 + nb_phases)].width = 12

    # Grand total for percentages
    all_montants = sum(cross.values())
    if all_montants == 0:
        all_montants = 1

    row = 4
    phase_totals = {p: 0 for p in phase_order}

    for lot in lot_order:
        row_vals = [lot]
        row_total = 0
        for phase in phase_order:
            val = cross.get((lot, phase), 0)
            row_vals.append(val)
            row_total += val
            phase_totals[phase] += val
        row_vals.append(row_total)
        row_vals.append(row_total / all_montants)

        for col, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col == len(row_vals):  # percentage
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col >= 2:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")

        row += 1

    # Column totals
    totals_vals = [_t("TOTAL", "TOTAL", lang)]
    col_grand = 0
    for phase in phase_order:
        totals_vals.append(phase_totals[phase])
        col_grand += phase_totals[phase]
    totals_vals.append(col_grand)
    totals_vals.append(1.0)

    _add_row(ws3, row, totals_vals, bold=True, subtotal=True)
    ws3.cell(row=row, column=len(totals_vals)).number_format = PCT_FMT

    # Column percentages row
    row += 1
    pct_vals = [_t("% par phase", "% per phase", lang)]
    for phase in phase_order:
        pct_vals.append(phase_totals[phase] / all_montants)
    pct_vals.append(1.0)
    pct_vals.append("")

    for col, val in enumerate(pct_vals, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = Font(name="Calibri", size=10, italic=True)
        cell.border = THIN_BORDER
        if isinstance(val, float):
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal="right")

    # -----------------------------------------------------------------------
    # Save and return bytes
    # -----------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
